# Mecanismo de Classificação Evolutiva — Modelo CredituS (Streamlit + Plotly)
# ------------------------------------------------------------
# - Algoritmo genético para classificar Adimplência (A) x Inadimplência (I)
# - Cromossomo = Eneadecaquatérnio Q (bias b0 + 18 genes)
# - Fitness FC = PA * PI (produto força acerto balanceado nas duas classes)
# - Seleção por roleta, crossover multi-ponto, mutação, substituição elitista
# - Geração de dados sintéticos no padrão CredituS OU upload de planilha (.xlsx)
# - Diagnósticos: convergência + matriz de confusão + distribuição de Q + roleta

import io
import time

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


# ----------------------------
# 0) CONFIGURAÇÃO DA PÁGINA (DEVE SER A PRIMEIRA)
# ----------------------------
st.set_page_config(page_title="Algoritmo Genético", layout="wide")


# ----------------------------
# 1) ESTILO (CSS)
# ----------------------------
st.markdown(
    """
<style>
:root {
  --bg: #09090b;
  --border: rgba(255, 255, 255, 0.05);
  --muted: #a1a1aa;
  --muted2: #52525b;
  --accent: #ef4444;
  --accent2: #3b82f6;
  --green: #10b981;
}

.main { background-color: var(--bg); }
section[data-testid="stSidebar"] { background-color: #0b1020; border-right: 1px solid var(--border); }
div[data-testid="stMetric"]{
  background: linear-gradient(180deg, rgba(255,255,255,0.03), rgba(255,255,255,0.01));
  border: 1px solid var(--border);
  border-radius: 12px;
  padding: 16px;
}

.small-muted { color: var(--muted); font-size: 0.92rem; margin-bottom: 1rem; }
.badge {
  display:inline-block; padding: 0.2rem 0.6rem; border-radius: 999px;
  background: rgba(255,255,255,0.05); border: 1px solid rgba(255,255,255,0.08);
  color: #d4d4d8; font-size: 0.82rem; margin-right: 0.4rem;
}
.footer { text-align:center; color: var(--muted2); margin-top: 3rem; font-size: 0.85rem; }

/* Destaque para a equação do classificador no topo */
.function-display {
    text-align: center;
    padding: 2rem 0;
}
</style>
""",
    unsafe_allow_html=True,
)


# ----------------------------
# 2) MOTOR GENÉTICO (ENGINE)
# ----------------------------
def criar_cromossomos(qtd_cromossomos: int, qtd_genes: int) -> np.ndarray:
    """Inicializa a população. Cada gene é sorteado no domínio [-1, +1]."""
    return -1 + 2 * np.random.rand(qtd_cromossomos, qtd_genes)


def calcular_fitness(
    cromossomos: np.ndarray,
    array_dados_clientes: np.ndarray,
    array_gabarito: np.ndarray,
) -> np.ndarray:
    total_adimplentes = np.sum(array_gabarito == 1)
    total_inadimplentes = np.sum(array_gabarito == 0)

    fitnesses = []
    for linha in cromossomos:
        bias = linha[0]
        genes = linha[1:]

        q = np.dot(array_dados_clientes, genes) + bias
        vetor_hipotese = np.where(q >= 0, 1, 0)

        acertos_adimplentes = np.sum((vetor_hipotese == 1) & (array_gabarito == 1))
        acertos_inadimplentes = np.sum((vetor_hipotese == 0) & (array_gabarito == 0))

        pa = acertos_adimplentes / total_adimplentes if total_adimplentes else 0.0
        pi = acertos_inadimplentes / total_inadimplentes if total_inadimplentes else 0.0

        fitnesses.append(pa * pi)

    return np.array(fitnesses, dtype=float)


def fitness_percentual(vetor_fitnesses: np.ndarray) -> np.ndarray:
    soma = np.sum(vetor_fitnesses)
    if soma == 0:
        return np.ones(len(vetor_fitnesses)) / len(vetor_fitnesses)
    return vetor_fitnesses / soma


def selecionar_pais_roleta(
    cromossomos: np.ndarray,
    percentual_fitnesses: np.ndarray,
) -> tuple[np.ndarray, np.ndarray]:
    roleta_acumulada = np.cumsum(percentual_fitnesses)
    indice_pai = min(int(np.searchsorted(roleta_acumulada, np.random.rand())), len(cromossomos) - 1)
    indice_mae = min(int(np.searchsorted(roleta_acumulada, np.random.rand())), len(cromossomos) - 1)
    return cromossomos[indice_pai], cromossomos[indice_mae]


def cruzar_pais(pai: np.ndarray, mae: np.ndarray) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    c1 = np.random.randint(1, len(pai))
    c2 = np.random.randint(1, len(pai))
    c3 = np.random.randint(1, len(pai))

    filho1 = np.concatenate([pai[:c1], mae[c1:]])
    filho2 = np.concatenate([pai[:c2], mae[c2:]])
    filho3 = np.concatenate([pai[:c3], mae[c3:]])

    return filho1, filho2, filho3


def mutar(
    filho1: np.ndarray, filho2: np.ndarray, filho3: np.ndarray
) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    for filho in [filho1, filho2, filho3]:
        indice = np.random.randint(0, len(filho))
        filho[indice] = -1 + 2 * np.random.rand()

    return filho1, filho2, filho3


def algoritmo_genetico(
    array_dados_clientes: np.ndarray,
    array_gabarito: np.ndarray,
    qtd_cromossomos: int = 12,
    geracoes: int = 500,
    fitness_alvo: float = 0.90,
    cruzamentos_por_geracao: int = 3,
    seed: int = 42,
) -> dict:
    np.random.seed(seed)
    qtd_genes = array_dados_clientes.shape[1] + 1
    populacao = criar_cromossomos(qtd_cromossomos, qtd_genes)
    fitnesses = calcular_fitness(populacao, array_dados_clientes, array_gabarito)

    melhor_cromossomo = populacao[int(np.argmax(fitnesses))].copy()
    melhor_fitness = float(np.max(fitnesses))

    hist_melhor = []
    hist_media = []
    geracao_convergencia = geracoes

    for geracao in range(geracoes):
        idx_melhor = int(np.argmax(fitnesses))
        if fitnesses[idx_melhor] > melhor_fitness:
            melhor_fitness = float(fitnesses[idx_melhor])
            melhor_cromossomo = populacao[idx_melhor].copy()

        hist_melhor.append(melhor_fitness)
        hist_media.append(float(np.mean(fitnesses)))

        if melhor_fitness >= fitness_alvo:
            geracao_convergencia = geracao + 1
            break

        percentuais = fitness_percentual(fitnesses)
        prole = []
        for _ in range(cruzamentos_por_geracao):
            pai, mae = selecionar_pais_roleta(populacao, percentuais)
            filho1, filho2, filho3 = cruzar_pais(pai, mae)
            filho1, filho2, filho3 = mutar(filho1, filho2, filho3)
            prole.extend([filho1, filho2, filho3])

        prole = np.array(prole)
        fitness_prole = calcular_fitness(prole, array_dados_clientes, array_gabarito)

        todos = np.vstack([populacao, prole])
        fitness_todos = np.concatenate([fitnesses, fitness_prole])
        ordem = np.argsort(fitness_todos)[::-1][:qtd_cromossomos]
        populacao = todos[ordem]
        fitnesses = fitness_todos[ordem]

    return {
        "melhor_cromossomo": melhor_cromossomo,
        "melhor_fitness": float(max(melhor_fitness, 0.0)),
        "hist_melhor": np.array(hist_melhor, dtype=float),
        "hist_media": np.array(hist_media, dtype=float),
        "geracao_convergencia": int(geracao_convergencia),
        "populacao_final": populacao,
        "fitness_final": fitnesses,
    }


def diagnosticar(
    cromossomo: np.ndarray,
    array_dados_clientes: np.ndarray,
    array_gabarito: np.ndarray,
) -> dict:
    bias = cromossomo[0]
    genes = cromossomo[1:]

    q = np.dot(array_dados_clientes, genes) + bias
    hipotese = np.where(q >= 0, 1, 0)

    ta = int(np.sum(array_gabarito == 1))
    ti = int(np.sum(array_gabarito == 0))
    nac = int(np.sum((hipotese == 1) & (array_gabarito == 1)))
    nic = int(np.sum((hipotese == 0) & (array_gabarito == 0)))

    pa = nac / ta if ta else 0.0
    pi = nic / ti if ti else 0.0

    return {
        "q": q,
        "hipotese": hipotese,
        "pa": pa,
        "pi": pi,
        "fc": pa * pi,
        "acuracia": (nac + nic) / len(array_gabarito),
        "nac": nac,
        "nic": nic,
        "ta": ta,
        "ti": ti,
        "fa": ta - nac,
        "fi": ti - nic,
    }


# ----------------------------
# 3) DADOS E UTILITÁRIOS DE CACHE
# ----------------------------
def nomes_colunas() -> list[str]:
    nomes = []
    for a in range(1, 10):
        nomes.append(f"X{a}1")
        nomes.append(f"X{a}2")
    return nomes


@st.cache_data(show_spinner=False)
def gerar_dados_sinteticos(n_clientes: int, ruido: float, seed: int) -> tuple[np.ndarray, np.ndarray]:
    rng = np.random.default_rng(seed)

    X = np.zeros((n_clientes, 18), dtype=int)
    for i in range(9):
        nuance1 = rng.integers(0, 2, size=n_clientes)
        X[:, 2 * i] = nuance1
        X[:, 2 * i + 1] = 1 - nuance1

    pesos_true = rng.uniform(-1, 1, size=18)
    bias_true = rng.uniform(-1, 1)
    q_true = X @ pesos_true + bias_true
    rotulos = (q_true >= 0).astype(int)

    if ruido > 0:
        flip = rng.random(n_clientes) < ruido
        rotulos[flip] = 1 - rotulos[flip]

    return X.astype(float), rotulos.astype(int)


def normalizar_gabarito(coluna: pd.Series) -> np.ndarray:
    if coluna.dtype == object:
        return coluna.map(
            lambda v: 1 if str(v).strip().upper() in {"A", "1", "ADIMPLENTE"} else 0
        ).to_numpy()
    return (coluna.to_numpy() >= 1).astype(int)


@st.cache_data(show_spinner=False)
def construir_template_xlsx() -> bytes:
    NAVY, LIGHT, GREEN, RED = "1F2A44", "EEF1F7", "DDF3E6", "F8E0E0"

    wb = Workbook()
    head_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    head_fill = PatternFill("solid", fgColor=NAVY)
    body_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="C9D2E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Aba Dados
    ws = wb.active
    ws.title = "Dados"
    cols_x = []
    for a in range(1, 10):
        cols_x += [f"X{a}1", f"X{a}2"]
    headers = ["ID"] + cols_x + ["Classe (A/I)"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill, cell.alignment, cell.border = head_font, head_fill, center, border

    exemplos = [
        [1, 1, 0, 0, 1, 1, 0, 0, 1, 0, 1, 1, 0, 0, 1, 1, 0, 1, 0, "A"],
        [2, 0, 1, 0, 1, 0, 1, 0, 1, 0, 1, 1, 0, 0, 1, 1, 0, 1, 0, "A"],
        [3, 1, 0, 1, 0, 0, 1, 1, 0, 1, 0, 0, 1, 1, 0, 0, 1, 0, 1, "I"],
        [4, 1, 0, 0, 1, 0, 1, 1, 0, 1, 0, 1, 0, 1, 0, 0, 1, 0, 1, "I"],
        [5, 0, 1, 1, 0, 0, 1, 1, 0, 1, 0, 0, 1, 1, 0, 0, 1, 0, 1, "I"],
    ]
    for row in exemplos:
        ws.append(row)
    for r in range(2, 2 + len(exemplos)):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font, cell.alignment, cell.border = body_font, center, border
            if c == len(headers):
                cell.fill = PatternFill("solid", fgColor=GREEN if cell.value == "A" else RED)

    dv01 = DataValidation(type="list", formula1='"0,1"', allow_blank=True,
                          showErrorMessage=True, error="Use apenas 0 ou 1.")
    ws.add_data_validation(dv01)
    dv01.add("B2:S1000")
    dvAI = DataValidation(type="list", formula1='"A,I"', allow_blank=True,
                          showErrorMessage=True, error="Use A (Adimplente) ou I (Inadimplente).")
    ws.add_data_validation(dvAI)
    dvAI.add("T2:T1000")

    ws.column_dimensions["A"].width = 6
    for i in range(18):
        ws.column_dimensions[chr(ord("B") + i)].width = 6
    ws.column_dimensions["T"].width = 14
    ws.freeze_panes = "A2"

    # Aba Legenda
    leg = wb.create_sheet("Legenda")
    leg.append(["Atributo", "Variavel", "Significado"])
    for c in range(1, 4):
        cell = leg.cell(row=1, column=c)
        cell.font, cell.fill, cell.alignment, cell.border = head_font, head_fill, center, border
    legenda = [
        ("Sexo", "X11", "Avaliado do sexo masculino"), ("Sexo", "X12", "Avaliado do sexo feminino"),
        ("Idade", "X21", "Idade do cliente <= 30 anos"), ("Idade", "X22", "Idade do cliente > 30 anos"),
        ("CEP", "X31", "CEP comercial (Faixa A)"), ("CEP", "X32", "CEP comercial (Faixa B)"),
        ("Emprego", "X41", "Tempo no emprego <= 2 anos"), ("Emprego", "X42", "Tempo no emprego > 2 anos"),
        ("Moradia", "X51", "Tempo de moradia <= 1 ano"), ("Moradia", "X52", "Tempo de moradia > 1 ano"),
        ("Uniao", "X61", "Estado civil (casado)"), ("Uniao", "X62", "Estado civil (solteiro)"),
        ("Atividade", "X71", "Trabalhador tec. mecanico"), ("Atividade", "X72", "Trabalhador tec. comercial"),
        ("Valor", "X81", "Valor do emprestimo <= 30 u.m"), ("Valor", "X82", "Valor do emprestimo > 30 u.m"),
        ("Parcelas", "X91", "Numero de parcelas <= 10"), ("Parcelas", "X92", "Numero de parcelas > 10"),
    ]
    for i, row in enumerate(legenda):
        leg.append(list(row))
        for c in range(1, 4):
            cell = leg.cell(row=i + 2, column=c)
            cell.font, cell.border = body_font, border
            cell.alignment = Alignment(vertical="center", horizontal="center" if c < 3 else "left")
            if (i // 2) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
    leg.column_dimensions["A"].width = 12
    leg.column_dimensions["B"].width = 10
    leg.column_dimensions["C"].width = 36
    leg.freeze_panes = "A2"

    # Aba Instruções
    ins = wb.create_sheet("Instrucoes")
    ins.column_dimensions["A"].width = 100
    linhas = [
        ("Como preencher o template CredituS", Font(name="Arial", bold=True, size=13, color=NAVY)),
        ("", body_font),
        ("1) Cada linha = um cliente do banco de dados.", body_font),
        ("2) Coluna ID: identificador livre. E ignorada pelo modelo.", body_font),
        ("3) Colunas X11..X92: variaveis binarias (0 ou 1). Veja a aba 'Legenda'.", body_font),
        ("   Regra das nuances: em cada par (Xa1, Xa2) exatamente UM vale 1 (one-hot).", body_font),
        ("   Ex.: homem -> X11=1, X12=0 ; mulher -> X11=0, X12=1.", body_font),
        ("4) Coluna Classe: A = Adimplente (pagou) ; I = Inadimplente (nao pagou).", body_font),
        ("5) Quanto mais clientes e mais equilibrio entre A e I, melhor o treino.", body_font),
        ("6) Salve em .xlsx e suba aqui na pagina 'Algoritmo Genetico'.", body_font),
        ("", body_font),
        ("Obs.: linhas incompletas (celulas vazias) sao descartadas no carregamento.", body_font),
    ]
    for i, (t, f) in enumerate(linhas):
        cell = ins.cell(row=i + 1, column=1, value=t)
        cell.font = f
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ----------------------------
# 4) PAINEL TEÓRICO
# ----------------------------
def theory_panel():
    st.markdown("### Teoria Explicada")
    st.markdown(
        "<span class='badge'>Eneadecaquatérnio</span> "
        "<span class='badge'>Fitness Produto</span> "
        "<span class='badge'>Seleção por Roleta</span> "
        "<span class='badge'>Crossover & Mutação</span>",
        unsafe_allow_html=True,
    )
    
    st.write("") # Whitespace em vez de <hr>

    with st.expander("Abrir teoria completa (cromossomo, fitness, roleta, evolução)", expanded=False):
        st.markdown(
            "Este motor resolve o problema **CredituS**: dado um banco de clientes descritos por "
            "atributos binários, evoluir um classificador que separe **Adimplentes (A)** de "
            "**Inadimplentes (I)** — aprendizado de máquina por seleção natural."
        )

        st.write("")

        st.markdown("### 1) O Cromossomo: Eneadecaquatérnio $Q$")
        st.markdown(
            "Cada cliente é descrito por 18 variáveis binárias (9 atributos $\\times$ 2 nuances). "
            "O classificador é o eneadecaquatérnio:"
        )
        st.latex(r"Q = b_0 + b_{11}X_{11} + b_{12}X_{12} + \cdots + b_{91}X_{91} + b_{92}X_{92}")
        st.markdown(
            "- $b_0$ é o **escalar** (bias); os 18 $b_{an}$ formam o **vetor** de pesos.\n"
            "- Um **cromossomo** é justamente esse conjunto de 19 genes "
            "$(b_0,\\, b_{11},\\, \\ldots,\\, b_{92})$, cada um sorteado no domínio $[-1,\\, +1]$.\n"
            "- A classificação é o **sinal** de $Q$, com *threshold* em zero:"
        )
        st.latex(r"Q \geq 0 \;\Rightarrow\; \text{Adimplente (A)} \qquad Q < 0 \;\Rightarrow\; \text{Inadimplente (I)}")
        st.markdown(
            "Antes do treino os genes são aleatórios (*Deixis am Phantasma*). "
            "Após a evolução, os genes ajustados formam um classificador útil (*Deixis ad Oculus*)."
        )

        st.write("")

        st.markdown("### 2) Fitness do Cromossomo")
        st.markdown(
            "A qualidade é medida combinando o acerto nas **duas** classes. Sejam NAC/NIC os "
            "adimplentes/inadimplentes classificados corretamente e TA/TI seus totais:"
        )
        st.latex(r"PA = \frac{NAC}{TA} \qquad\qquad PI = \frac{NIC}{TI}")
        st.markdown("O fitness é o **produto**:")
        st.latex(r"FC = PA \cdot PI = \frac{NAC}{TA}\cdot\frac{NIC}{TI}")
        st.markdown(
            "O produto é proposital: um modelo que chuta *todos adimplentes* acerta $PA = 1$ "
            "mas $PI = 0$, logo $FC = 0$. Só obtém fitness alto quem acerta as duas classes — "
            "é o que mata o classificador preguiçoso."
        )

        st.write("")

        st.markdown("### 3) Seleção por Roleta")
        st.markdown("A chance de um cromossomo ser pai é proporcional ao seu fitness:")
        st.latex(r"p_i = \frac{FC_i}{\sum_k FC_k}")
        st.markdown(
            "Monta-se uma roleta (soma acumulada) e sorteia-se um número em $[0,\\,1]$. "
            "Quem tem mais fitness ocupa fatia maior e é escolhido com mais frequência."
        )

        st.write("")

        st.markdown("### 4) Crossover (Cruzamento)")
        st.markdown(
            "Dois pais trocam segmentos em pontos de corte sorteados, gerando filhos que "
            "herdam genes de ambos. Aqui são gerados **três filhos** a partir de três cortes."
        )

        st.markdown("### 5) Mutação")
        st.markdown(
            "Em cada filho, um gene é sorteado e substituído por um novo valor aleatório em "
            "$[-1,\\,+1]$. A mutação injeta diversidade e evita ficar preso em ótimos locais."
        )

        st.markdown("### 6) Nova Linhagem (Substituição Elitista)")
        st.markdown(
            "Os **dois piores** cromossomos da população são substituídos pelos **dois melhores** "
            "filhos da iteração. Os bons sobreviventes são preservados (elitismo), de modo que o "
            "melhor fitness nunca regride."
        )

        st.write("")

        st.markdown("### 7) Convergência")
        st.markdown(
            "A cada geração o melhor fitness tende a subir, aproximando-se de 1. Quando atinge o "
            "alvo definido, o treino encerra e o cromossomo campeão vira o classificador final:"
        )
        st.latex(r"FC \to 1 \quad \Longleftrightarrow \quad PA \to 1 \;\text{ e }\; PI \to 1")


# ----------------------------
# 5) GRÁFICOS
# ----------------------------
def make_convergence_plot(hist_melhor: np.ndarray, hist_media: np.ndarray) -> go.Figure:
    geracoes = np.arange(1, len(hist_melhor) + 1)

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=geracoes, y=hist_media, mode="lines",
        name="Fitness médio da população",
        line=dict(color="#3b82f6", width=2, dash="dot"),
        hovertemplate="geração=%{x}<br>média=%{y:.4f}<extra></extra>",
    ))
    fig.add_trace(go.Scatter(
        x=geracoes, y=hist_melhor, mode="lines",
        name="Melhor fitness (elite)",
        line=dict(color="#ef4444", width=3),
        hovertemplate="geração=%{x}<br>melhor=%{y:.4f}<extra></extra>",
    ))

    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=20, r=20, t=60, b=20),
        title="Evolução do Fitness por Geração",
        xaxis_title="Geração",
        yaxis_title="Fitness (FC = PA · PI)",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        transition=dict(duration=450),
    )
    fig.update_xaxes(showgrid=True, gridcolor="rgba(255,255,255,0.06)")
    fig.update_yaxes(showgrid=True, gridcolor="rgba(255,255,255,0.06)", range=[0, 1.02])
    return fig


def make_chromosome_plot(cromossomo: np.ndarray, nomes: list[str]) -> go.Figure:
    rotulos = ["b<sub>0</sub>"]
    for a in range(1, 10):
        rotulos.append(f"b<sub>{a}1</sub>")
        rotulos.append(f"b<sub>{a}2</sub>")
    rotulos = rotulos[:len(cromossomo)]

    cores = ["#10b981" if v >= 0 else "#ef4444" for v in cromossomo]

    fig = go.Figure(go.Bar(
        x=rotulos, y=cromossomo,
        marker=dict(color=cores, line=dict(color="rgba(255,255,255,0.25)", width=0.5)),
        hovertemplate="gene=%{x}<br>peso=%{y:.4f}<extra></extra>",
    ))
    fig.add_hline(y=0, line_width=1, line_color="rgba(229,231,235,0.45)")

    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=20, r=20, t=60, b=20),
        title="Melhor Cromossomo — Eneadecaquatérnio Treinado (Deixis ad Oculus)",
        xaxis_title="Genes",
        yaxis_title="Valor do gene",
        transition=dict(duration=450),
    )
    fig.update_xaxes(showgrid=False)
    fig.update_yaxes(showgrid=True, gridcolor="rgba(255,255,255,0.06)")
    return fig


def make_q_distribution_plot(q: np.ndarray, hipotese: np.ndarray) -> go.Figure:
    fig = go.Figure()
    fig.add_trace(go.Histogram(
        x=q[hipotese == 1], name="Classificado A (Q ≥ 0)",
        marker_color="#10b981", opacity=0.7,
    ))
    fig.add_trace(go.Histogram(
        x=q[hipotese == 0], name="Classificado I (Q < 0)",
        marker_color="#ef4444", opacity=0.7,
    ))
    fig.add_vline(
        x=0, line_width=2, line_dash="dash", line_color="rgba(229,231,235,0.85)",
        annotation_text="Threshold = 0", annotation_position="top",
    )

    fig.update_layout(
        barmode="overlay",
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=20, r=20, t=60, b=20),
        title="Distribuição dos Valores de Q (Saída do Classificador)",
        xaxis_title="Q (eneadecaquatérnio)",
        yaxis_title="Nº de clientes",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    )
    fig.update_xaxes(showgrid=True, gridcolor="rgba(255,255,255,0.06)")
    fig.update_yaxes(showgrid=True, gridcolor="rgba(255,255,255,0.06)")
    return fig


def make_confusion_plot(diag: dict) -> go.Figure:
    z = [
        [diag["nac"], diag["fa"]],
        [diag["fi"], diag["nic"]],
    ]
    texto = [
        [f"NAC<br>{diag['nac']}", f"erro<br>{diag['fa']}"],
        [f"erro<br>{diag['fi']}", f"NIC<br>{diag['nic']}"],
    ]

    fig = go.Figure(go.Heatmap(
        z=z,
        x=["Predito A", "Predito I"],
        y=["Real A", "Real I"],
        text=texto,
        texttemplate="%{text}",
        colorscale=[[0, "#09090b"], [1, "#ef4444"]],
        showscale=False,
        hovertemplate="%{y} / %{x}: %{z}<extra></extra>",
    ))
    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=20, r=20, t=60, b=20),
        title="Matriz de Confusão (NAC e NIC na diagonal)",
    )
    return fig


def make_roleta_plot(fitness_final: np.ndarray) -> go.Figure:
    pct = fitness_percentual(fitness_final)
    rotulos = [f"C{i + 1}" for i in range(len(fitness_final))]

    fig = go.Figure(go.Pie(
        labels=rotulos,
        values=pct,
        hole=0.45,
        textinfo="label+percent",
        marker=dict(line=dict(color="#09090b", width=2)),
        hovertemplate="%{label}<br>fatia=%{percent}<extra></extra>",
    ))
    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=20, r=20, t=60, b=20),
        title="Roda do Acaso — Probabilidade de Seleção",
        showlegend=False,
    )
    return fig


# ----------------------------
# 6) CABEÇALHO
# ----------------------------
st.title("Algoritmo Genético")
st.caption("Classificador Evolutivo CredituS — Adimplência (A) × Inadimplência (I)")
st.write("")

theory_panel()
st.write("")


# ----------------------------
# 7) SIDEBAR (BARRA LATERAL)
# ----------------------------
st.sidebar.header("Controles")

st.sidebar.download_button(
    "⬇ Baixar template (.xlsx)",
    data=construir_template_xlsx(),
    file_name="CredituS_template.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
st.sidebar.caption("Preencha o modelo e suba abaixo.")

st.sidebar.divider()

fonte = st.sidebar.radio(
    "Fonte de dados",
    ["Gerar dados sintéticos (padrão CredituS)", "Enviar planilha (.xlsx)"],
    index=0,
)

dados_ok = False
features = None
gabarito = None
nomes = nomes_colunas()

if fonte.startswith("Enviar"):
    up = st.sidebar.file_uploader("Planilha de clientes", type=["xlsx", "xls"])
    st.sidebar.caption("1ª coluna = índice (ignorada) • última coluna = gabarito (A / I).")
    if up is not None:
        try:
            df = pd.read_excel(up)
            df_features = df.iloc[:, 1:-1].apply(pd.to_numeric, errors="coerce")
            serie_gab = df.iloc[:, -1]
            
            mascara = df_features.notna().all(axis=1) & serie_gab.notna()
            n_descartadas = int((~mascara).sum())
            df_features = df_features[mascara]
            serie_gab = serie_gab[mascara]

            features = df_features.to_numpy(dtype=float)
            gabarito = normalizar_gabarito(serie_gab)
            nomes = list(df.iloc[:, 1:-1].columns.astype(str))
            dados_ok = len(gabarito) > 0
            
            if n_descartadas:
                st.sidebar.warning(f"{n_descartadas} linha(s) incompleta(s) descartada(s).")
            if not dados_ok:
                st.sidebar.error("Nenhuma linha completa encontrada na planilha.")
        except Exception as e:
            st.sidebar.error(f"Falha ao ler a planilha: {e}")
else:
    n_clientes = st.sidebar.slider("Nº de clientes", 50, 1000, 300, step=50)
    ruido = st.sidebar.slider("Ruído dos rótulos", 0.0, 0.30, 0.05, step=0.01)
    seed_dados = st.sidebar.number_input("Seed dos dados", value=7, step=1)
    features, gabarito = gerar_dados_sinteticos(n_clientes, ruido, int(seed_dados))
    dados_ok = True

st.sidebar.divider()
st.sidebar.subheader("Evolução")

geracoes = st.sidebar.slider("Máx. de gerações", 50, 2000, 500, step=50)
fitness_alvo = st.sidebar.slider("Fitness alvo", 0.50, 0.99, 0.85, step=0.01)

with st.sidebar.expander("⚙️ Parâmetros Avançados (Motor Genético)"):
    qtd_cromossomos = st.slider("Tamanho da população", 4, 50, 12, step=1)
    cruzamentos = st.slider("Cruzamentos por geração", 1, 20, 3, step=1)
    seed_ga = st.number_input("Seed da evolução", value=42, step=1)
    st.caption("Cada cruzamento gera 3 filhos. Mais cruzamentos → convergência rápida.")

st.sidebar.divider()
rodar = st.sidebar.button("▶ Rodar evolução", use_container_width=True, type="primary")


# ----------------------------
# 8) PRÉVIA DOS DADOS
# ----------------------------
if not dados_ok:
    st.info("Envie uma planilha na barra lateral para começar (ou troque para dados sintéticos).")
    st.markdown("<div class='footer'>The Everything Calculator - Fellipe Almässy • </div>", unsafe_allow_html=True)
    st.stop()

ta_total = int(np.sum(gabarito == 1))
ti_total = int(np.sum(gabarito == 0))

st.markdown(
    f"<div class='small-muted'>Banco carregado: <b>{len(gabarito)}</b> clientes • "
    f"<b>{features.shape[1]}</b> variáveis • "
    f"<b>{ta_total}</b> adimplentes (A) • <b>{ti_total}</b> inadimplentes (I).</div>",
    unsafe_allow_html=True,
)

df_preview = pd.DataFrame(features.astype(int), columns=nomes[:features.shape[1]])
df_preview.insert(0, "Classe Real", np.where(gabarito == 1, "A", "I"))
with st.expander("Ver amostra do banco de dados", expanded=False):
    st.dataframe(df_preview.head(15), use_container_width=True, hide_index=True)

st.write("")


# ----------------------------
# 9) RODAR ALGORITMO (com estado de sessão)
# ----------------------------
if "ga_result" not in st.session_state:
    st.session_state.ga_result = None

if rodar:
    with st.spinner("Evoluindo a população..."):
        t0 = time.time()
        res = algoritmo_genetico(
            features, gabarito,
            qtd_cromossomos=qtd_cromossomos,
            geracoes=geracoes,
            fitness_alvo=fitness_alvo,
            cruzamentos_por_geracao=cruzamentos,
            seed=int(seed_ga),
        )
        res["runtime"] = time.time() - t0
        res["diag"] = diagnosticar(res["melhor_cromossomo"], features, gabarito)
    st.session_state.ga_result = res
    st.toast(f"Evolução concluída em {res['runtime']:.2f}s!", icon="✅")

res = st.session_state.ga_result
if res is None:
    st.info("Configure os parâmetros e clique em **Rodar evolução** na barra lateral.")
    st.markdown("<div class='footer'>The Everything Calculator - Fellipe Almässy • </div>", unsafe_allow_html=True)
    st.stop()

diag = res["diag"]


# ----------------------------
# 10) MÉTRICAS
# ----------------------------
st.markdown("<div class='function-display'>", unsafe_allow_html=True)
st.latex(rf"FC_{{\text{{melhor}}}} = {res['melhor_fitness']:.4f}")
st.markdown("</div>", unsafe_allow_html=True)

m1, m2, m3, m4, m5 = st.columns([1.2, 1.0, 1.0, 1.1, 1.1])

m1.metric("Melhor Fitness", f"{res['melhor_fitness']:.4f}", "FC = PA · PI")
m2.metric("PA (acerto A)", f"{diag['pa'] * 100:.1f}%", f"{diag['nac']}/{diag['ta']}")
m3.metric("PI (acerto I)", f"{diag['pi'] * 100:.1f}%", f"{diag['nic']}/{diag['ti']}")
m4.metric("Acurácia Global", f"{diag['acuracia'] * 100:.1f}%", f"{diag['nac'] + diag['nic']}/{len(gabarito)}")
m5.metric("Convergiu em", f"{res['geracao_convergencia']} ger.", f"{res['runtime']:.2f}s")

st.write("")


# ----------------------------
# 11) ABAS: MOTOR / DIAGNÓSTICOS
# ----------------------------
tab_motor, tab_diag = st.tabs(["Visão do Motor", "Diagnósticos"])

with tab_motor:
    fig_conv = make_convergence_plot(res["hist_melhor"], res["hist_media"])
    st.plotly_chart(fig_conv, use_container_width=True, config={'displayModeBar': False})

    fig_chromo = make_chromosome_plot(res["melhor_cromossomo"], nomes)
    st.plotly_chart(fig_chromo, use_container_width=True, config={'displayModeBar': False})
    st.caption(
        "Barras verdes = peso positivo (empurra para Adimplente), "
        "vermelhas = peso negativo (empurra para Inadimplente)."
    )

    rotulos_genes = ["b0 (bias)"] + [f"b{a}{n}" for a in range(1, 10) for n in (1, 2)]
    df_chromo = pd.DataFrame({
        "gene": rotulos_genes[:len(res["melhor_cromossomo"])],
        "peso": res["melhor_cromossomo"],
    })
    st.download_button(
        "⬇ Baixar cromossomo treinado (.csv)",
        data=df_chromo.to_csv(index=False).encode("utf-8"),
        file_name="cromossomo_treinado.csv",
        mime="text/csv",
    )

with tab_diag:
    c1, c2 = st.columns(2)
    with c1:
        st.plotly_chart(make_confusion_plot(diag), use_container_width=True, config={'displayModeBar': False})
    with c2:
        st.plotly_chart(make_roleta_plot(res["fitness_final"]), use_container_width=True, config={'displayModeBar': False})

    st.plotly_chart(make_q_distribution_plot(diag["q"], diag["hipotese"]), use_container_width=True, config={'displayModeBar': False})

    st.markdown("<div class='small-muted'>Clientes com classificação do modelo:</div>", unsafe_allow_html=True)
    df_result = pd.DataFrame({
        "Classe Real": np.where(gabarito == 1, "A", "I"),
        "Q": np.round(diag["q"], 4),
        "Predito": np.where(diag["hipotese"] == 1, "A", "I"),
        "Acertou?": np.where(diag["hipotese"] == gabarito, "OK", "X"),
    })
    st.dataframe(df_result.head(25), use_container_width=True, hide_index=True)


# ----------------------------
# 12) RODAPÉ
# ----------------------------
st.markdown("<div class='footer'>The Everything Calculator - Fellipe Almässy • </div>", unsafe_allow_html=True)
